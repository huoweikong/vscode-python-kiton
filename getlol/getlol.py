import requests
import re
import openpyxl
import pymysql

def get_html(hero):
    headers = {
        'Referer': 'http://lol.qq.com/web201310/info-defail.shtml?id=' + hero,
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.106 Safari/537.36',
    }
    url = 'http://lol.qq.com/biz/hero/' + hero + '.js'
    # print(url)
    response = requests.get(url, headers=headers)
    response = response.text
    return response

def hero_info(response):
    # 英雄名称
    hero_name = re.findall(r'"name":"(.*?)","title"', response, re.S)[0]
    hero_title = re.findall(r'"title":"(.*?)","tags"', response, re.S)[0]
    # 技能(QWER)
    hero_spells = re.findall(r'"spells":(.*?),"passive"', response, re.S)[0]
    # 技能名称
    hero_spells_name = re.findall(
        r'"name":"(.*?)","description"', hero_spells, re.S)
    # 技能描述
    hero_spells_description = re.findall(
        r'"description":"(.*?)","image"', hero_spells, re.S)
    # 技能消耗
    hero_spells_resource = re.findall(
        r'"resource":"(.*?)"}', hero_spells, re.S)
    # 技能主被动
    hero_spells_group = re.findall(r'"group":"(.*?)","x"', hero_spells, re.S)
    spells_Q = hero_spells_name[0] + ':' + hero_spells_description[0] + \
        '|' + hero_spells_resource[0] + '|' + hero_spells_group[0]
    spells_W = hero_spells_name[1] + ':' + hero_spells_description[1] + \
        '|' + hero_spells_resource[1] + '|' + hero_spells_group[1]
    spells_E = hero_spells_name[2] + ':' + hero_spells_description[2] + \
        '|' + hero_spells_resource[2] + '|' + hero_spells_group[2]
    spells_R = hero_spells_name[3] + ':' + hero_spells_description[3] + \
        '|' + hero_spells_resource[3] + '|' + hero_spells_group[3]
    Spells = spells_Q + '\n' + spells_W + '\n' + spells_E + '\n' + spells_R
    # 被动技能
    hero_passive = re.findall(r'"passive":(.*?),"lore"', response, re.S)[0]
    # 被动技能名称
    hero_passive_name = re.findall(
        r'"name":"(.*?)","description"', hero_passive, re.S)[0]
    # 技能描述
    hero_passive_description = re.findall(
        r'"description":"(.*?)","image"', hero_passive, re.S)[0]
    # 技能主被动
    hero_passive_group = re.findall(
        r'"group":"(.*?)","x"', hero_passive, re.S)[0]
    passive = hero_passive_name + ':' + \
        hero_passive_description + '|' + hero_passive_group
    hero_spells_info = [hero_name, hero_title, passive, Spells]
    return hero_spells_info

def get_hero():
    with open('hero', 'r') as f:
        hero = f.readlines()
    return hero

def save_to_excel(her):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = '英雄称号'
    ws['B1'] = '英雄名称'
    ws['C1'] = '被动技能'
    ws['D1'] = '主动技能'
    for hero in her:
        ws.append(hero)
    wb.save('herotest.xlsx')

def save_to_mysql(her):
    for i in her:
        ch = '"' + i[0] + '"'
        name = '"' + i[1] + '"'
        bd_name = '"' + i[2] + '"'
        zd_name = '"' + i[3] + '"'
        db = pymysql.connect(host='localhost', user='root',
                             password='123456', database='python_mysql', charset='utf8')
        cursor = db.cursor()
        sql = ''' insert into lolheroinfo values (%s, %s, %s, %s);
        ''' % (ch, name, bd_name, zd_name)
        # print(sql)
        try:
            # 执行sql语句
            cursor.execute(sql)
            # 提交到数据库执行
            db.commit()
            print(ch, ' insert into success!')
        except:
            db.rollback()

        db.close()
    return True

def main():
    heros = get_hero()
    her = []
    for hero in heros:
        hero = hero.split('"')[3]
        response = get_html(hero)
        her_infos = hero_info(response)
        her_encode = []
        for i in her_infos:
            i = i.encode("latin-1").decode("unicode_escape")
            her_encode.append(i)
        her.append(her_encode)

    save_to_excel(her)
    #save_to_mysql(her)

if __name__ == '__main__':
    main()